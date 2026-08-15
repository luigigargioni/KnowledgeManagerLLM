# results_excel.py

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config_loader import RESULTS_DIR

RESULTS_EXCEL_PATH = RESULTS_DIR / "all_results.xlsx"

# Columns of the main sheet (one row per objective)
_OBJECTIVE_COLUMNS = [
    "test_date",
    "batch_id",
    "scenario_id",
    "patient",
    "overall_status",
    "turns",
    "elapsed_seconds",
    "objectives",
    "judge_check",
    "conversation",
    "initial_therapy",
    "final_therapy",
]

# Colors for the various statuses
_STATUS_FILLS = {
    "completed": PatternFill("solid", fgColor="C6EFCE"),  # green
    "partial": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "failed": PatternFill("solid", fgColor="FFC7CE"),  # red
    "not_attempted": PatternFill("solid", fgColor="D9D9D9"),  # gray
    "error": PatternFill("solid", fgColor="F4CCCC"),  # dark red
}

_HEADER_FILL = PatternFill("solid", fgColor="2F4F7F")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_CELL_FONT = Font(name="Arial", size=10)
_BORDER_SIDE = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(
    left=_BORDER_SIDE,
    right=_BORDER_SIDE,
    top=_BORDER_SIDE,
    bottom=_BORDER_SIDE,
)


def _get_or_create_workbook(path: Path = RESULTS_EXCEL_PATH) -> tuple[openpyxl.Workbook, bool]:
    """
    Load the existing workbook or create a new one.
    Returns (workbook, is_new).
    """
    if path.exists():
        return openpyxl.load_workbook(path), False

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the empty default sheet
    return wb, True


def _get_or_create_sheet(wb: openpyxl.Workbook, sheet_name: str, is_new_wb: bool):
    """
    Return the existing sheet or create a new one with headers.
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    _write_headers(ws, _OBJECTIVE_COLUMNS)
    return ws


def _write_header_cell(ws, col_idx: int, col_name: str) -> None:
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_headers(ws, columns: list[str]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        _write_header_cell(ws, col_idx, col_name)


def _sync_headers(ws) -> list[str]:
    """
    Return the sheet's actual column layout, appending any column this module
    knows about that the sheet does not have yet.

    Workbooks written by an earlier version lack the newer columns. Rewriting
    the header in the canonical order would shift every row already stored, so
    a missing column is added at the end instead and rows are always written by
    column name, never by position.
    """
    existing = [cell.value for cell in ws[1] if cell.value]
    if not existing:
        _write_headers(ws, _OBJECTIVE_COLUMNS)
        return list(_OBJECTIVE_COLUMNS)

    for col_name in _OBJECTIVE_COLUMNS:
        if col_name not in existing:
            existing.append(col_name)
            _write_header_cell(ws, len(existing), col_name)
    return existing


def _append_row(ws, columns: list[str], values: dict, status: str | None) -> int:
    """Append one row, matching values to the sheet's columns by name."""
    row_idx = ws.max_row + 1
    for col_idx, col_name in enumerate(columns, start=1):
        _style_cell(ws.cell(row=row_idx, column=col_idx, value=values.get(col_name, "")), status)
    return row_idx


def _style_cell(cell, status: str | None = None) -> None:
    cell.font = _CELL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if status and status in _STATUS_FILLS:
        cell.fill = _STATUS_FILLS[status]


def _set_column_widths(ws, columns: list[str]) -> None:
    widths = {
        "test_date": 18,
        "batch_id": 20,
        "scenario_id": 12,
        "patient": 22,
        "overall_status": 16,
        "turns": 8,
        "elapsed_seconds": 16,
        "judge_check": 50,
        "objectives": 50,
        "conversation": 80,
        "initial_therapy": 50,
        "final_therapy": 50,
    }
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)


def append_batch_results(
    evaluation: dict,
    batch_id: str,
    scenario: str,
    conversation: str,
    initial_therapy: dict,
    final_therapy: dict,
    excel_path: Path = RESULTS_EXCEL_PATH,
) -> Path:
    """
    Append the results of a batch run to the global Excel file.
    If the file does not exist it is created. If it exists, rows are
    appended preserving data from previous runs.

    Each objective of each scenario occupies a separate row.
    The scenario_id and test_date columns uniquely identify the run.

    Args:
        evaluations: list of dicts produced by JudgeAgent.evaluate(),
                     enriched by test.py with scenario_id, patient, turns, elapsed_seconds
        batch_id:    batch identifier (e.g. "20260630_143000")
        initial_therapy: the scenario's therapy as installed before the conversation
        final_therapy:   the therapy as it stands at the end of the conversation
        excel_path:  path of the global Excel file (default: logs/batch_results/all_results.xlsx)

    Returns:
        Path of the updated Excel file.
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    wb, is_new_wb = _get_or_create_workbook(excel_path)
    ws = _get_or_create_sheet(wb, "Results", is_new_wb)
    columns = _sync_headers(ws)

    if ws.max_row == 1:
        _set_column_widths(ws, columns)

    test_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    values = {
        "test_date": test_date,
        "batch_id": batch_id,
        "scenario_id": evaluation.get("scenario_id", ""),
        "patient": evaluation.get("patient", ""),
        "turns": evaluation.get("turns", ""),
        "elapsed_seconds": evaluation.get("elapsed_seconds", ""),
        "objectives": scenario,
        "conversation": conversation or "",
        "initial_therapy": json.dumps(initial_therapy, indent=2),
    }

    if evaluation.get("status") == "error":
        # Failed scenario: write a single row with the error status. There is no
        # judge verdict, so the failure message and the judge's unparsed output
        # take the place of the checks.
        status = "error"
        raw_output = evaluation.get("raw_output", "")[:500]
        values |= {
            "overall_status": "error",
            "judge_check": "\n\n".join(
                filter(None, [evaluation.get("message", "Evaluation failed"), raw_output])
            ),
            "final_therapy": "",
        }
    else:
        status = evaluation.get("overall_status", "")
        values |= {
            "overall_status": status,
            "judge_check": json.dumps(evaluation.get("objectives", []), indent=2),
            "final_therapy": json.dumps(final_therapy, indent=2),
        }

    _append_row(ws, columns, values, status)

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    return excel_path
